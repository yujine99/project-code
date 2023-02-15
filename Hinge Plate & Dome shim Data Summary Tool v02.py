# -*- coding: utf-8 -*-
# @Time     : 2023/2/7 17:01
# @Author   : Carolyn_yu
# @Email    : Carolyn_yu@pegatroncorp.com
# @File     : Hinge Plate & Dome shim Data Summary.py
# @Software : PyCharm
# @Describe : ---------------------------
# 230210 First release
# 230210  02 修改代码格式问题,删除类后的（），

from pathlib import Path
from tkinter import Tk, Label, StringVar, Entry, Button, Listbox, Scrollbar, X, Y, RIGHT, BOTTOM, HORIZONTAL
# from tkinter.filedialog import askdirectory
from tkinter import filedialog, messagebox
from os import walk
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

source_item_n = ['hv0', 'hv0.05', 'hv0.10', 'hv0.15', 'hv0.20', 'hv0.25', 'hv0.30', 'hv0.35', 'hv0.40',
                 'dv0.07', 'dv0.12', 'dv0.17', 'dv0.22', 'dv0.27', 'dv0.32', 'dv0.37', 'dv0.425', 'dv0.48', 'dv0.53']

cell_s_ = 4  # 数据行前其它行个数
h_long_ = 9  # HINGE PLATE （metal） data个数
d_long_ = 10  # Dome shim（metal） data个数

m_item_dict_ = {}
m_item_n_list_ = []
f_item_dict_ = {}
f_item_n_list_ = []


def ck_model(filepath):
    with open(filepath, 'r') as f:
        data = f.read()
        m_model = '-M-'
        f_model = '-F-'
        # if data.find('-M-') >= 0:  # 判断是金属否-M- 方法一
        if m_model in data:  # 判断是包含str 方法二
            print(f"Model is {m_model}")
            return m_model
        elif f_model in data:
            print(f"Model is {f_model}")
            return f_model
        else:
            print(f"Model is error.")


def item_sum(value_l, m_str, *parame):
    if m_str == 'f':  # fabric小数点后不带0
        for j in parame:
            for i in j:
                i = m_str + i
                if len(i) >= 7:
                    # print(i)
                    if i[-1] == '0':
                        item_n = i[:len(i) - 1]
                        item_v = value_l.count(str(i[3:len(i) - 1]))
                        f_item_dict_[item_n] = item_v
                        f_item_n_list_.append(item_n)
                        # print(item_dict)
                        # print(item_n_list)
                        # print(item_dict['fhv0.1'])
                    else:
                        item_n = i
                        item_v = value_l.count(str(i[3:]))
                        f_item_dict_[item_n] = item_v
                        f_item_n_list_.append(item_n)
                else:
                    if i[-2:] == '.0':
                        item_n = i[:-2]
                        item_v = value_l.count(str(i[3:-2]))
                        f_item_dict_[item_n] = item_v
                        f_item_n_list_.append(item_n)
                    else:
                        item_n = i
                        item_v = value_l.count(str(i[3:]))
                        f_item_dict_[item_n] = item_v
                        f_item_n_list_.append(item_n)
    elif m_str == 'm':  # metal小数点后带0
        for j in parame:
            for i in j:
                i = m_str + i
                item_n = i
                item_v = value_l.count(str(i[3:]))
                m_item_dict_[item_n] = item_v
                m_item_n_list_.append(item_n)
    return f_item_n_list_, f_item_dict_, m_item_n_list_, m_item_dict_


def write_data(m_item_dict, f_item_dict, save_filename, cell_s, h_long, d_long):
    # cell_s = 4  # 数据行前其它行
    # h_long = 9  # HINGE PLATE （metal） data个数
    # d_long = 10  # Dome shim（metal） data个数
    wb = openpyxl.Workbook()
    wb.create_sheet(index=0, title="Sheet1")
    sheet = wb['Sheet1']

    title = [['HINGE PLATE （metal）', '', '', '', '', 'HINGE PLATE （fabric）', '', '', '', '', 'metal+fabric'],
             ['规格', '数量', 'total ', '比例', '', '规格', '数量', 'total ', '比例', '', '总比例']]
    # 写入表格栏名
    for i in range(0, len(title)):
        for j in range(0, len(title[i])):
            sheet.cell(row=i + 3, column=j + 1, value=str(title[i][j]))
            sheet.cell(row=i + 3 + h_long + 5, column=j + 1, value=str(title[i][j]))

    # 合并单元格
    sheet.merge_cells('A3:D3')
    sheet.merge_cells('F3:I3')
    sheet.merge_cells('A{}:D{}'.format(cell_s + h_long + 4, cell_s + h_long + 4))
    sheet.merge_cells('F{}:I{}'.format(cell_s + h_long + 4, cell_s + h_long + 4))

    '''写入 metal data ROW A-D'''
    cell_no = cell_s
    for key, value in m_item_dict.items():
        cell_no += 1
        # HINGE PLATE （metal） data row5~13
        if cell_s <= cell_no <= cell_s + h_long:
            sheet['A' + str(cell_no)] = key[3:]
            sheet['B' + str(cell_no)] = value
            sheet['D' + str(cell_no)] = '=B{}/C{}'.format(cell_no, cell_s + 1)  # 比例=B5/C5
            sheet['D' + str(cell_no)].number_format = '0.00%'
            sheet['K' + str(cell_no)] = '=(B{}+G{})/(C{}+H{})'.format(cell_no, cell_no, cell_s + 1,
                                                                      cell_s + 1)  # 总比例=(B5+G5)/(C5+H5)
            sheet['K' + str(cell_no)].number_format = '0.00%'
        # Dome shim（metal） row19~28
        else:
            sheet['A' + str(cell_no + 5)] = key[3:]
            sheet['B' + str(cell_no + 5)] = value
            sheet['D' + str(cell_no + 5)] = '=B{}/C{}'.format(cell_no + 5, cell_s + 1 + h_long + 5)  # 比例=B19/C19
            sheet['D' + str(cell_no + 5)].number_format = '0.00%'
            # 总比例=(B19+G19)/(C19+H19)
            sheet['K' + str(cell_no + 5)] = '=(B{}+G{})/(C{}+H{})'.format(cell_no + 5, cell_no + 5,
                                                                          cell_s + 1 + h_long + 5,
                                                                          cell_s + 1 + h_long + 5)
            sheet['K' + str(cell_no + 5)].number_format = '0.00%'
    # sheet['C5'] = '=SUM(B5:B13)'  # 求HINGE 和 方法一
    # 求HINGE PLATE （metal）总数和 方法二
    sheet['C{}'.format(cell_s + 1)] = '=SUM(B{}:B{})'.format(cell_s + 1, cell_s + h_long)
    # 求Dome shim（metal）总数和
    sheet['C{}'.format(cell_s + 1 + h_long + 5)] = '=SUM(B{}:B{})'.format(cell_s + 1 + h_long + 5,
                                                                          cell_s + h_long + 5 + d_long)

    '''写入 fabric data ROW F-G'''
    cell_no = cell_s
    for key, value in f_item_dict.items():
        cell_no += 1
        # HINGE PLATE （fabric） data row5~13
        if cell_s <= cell_no <= cell_s + h_long:
            sheet['F' + str(cell_no)] = key[3:]
            sheet['G' + str(cell_no)] = value
            sheet['I' + str(cell_no)] = '=G{}/H{}'.format(cell_no, cell_s + 1)  # 比例=G5/H5
            sheet['I' + str(cell_no)].number_format = '0.00%'
        # Dome shim（fabric） row19~28
        else:
            sheet['F' + str(cell_no + 5)] = key[3:]
            sheet['G' + str(cell_no + 5)] = value
            sheet['I' + str(cell_no + 5)] = '=G{}/H{}'.format(cell_no + 5, cell_s + 1 + h_long + 5)  # 比例=G19/H19
            sheet['I' + str(cell_no + 5)].number_format = '0.00%'
    # 求HINGE PLATE （fabric）总数和
    sheet['H{}'.format(cell_s + 1)] = '=SUM(G{}:G{})'.format(cell_s + 1, cell_s + h_long)
    # 求Dome shim（fabric）总数和
    sheet['H{}'.format(cell_s + 1 + h_long + 5)] = '=SUM(G{}:G{})'.format(cell_s + 1 + h_long + 5,
                                                                          cell_s + h_long + 5 + d_long)

    wb.save(save_filename)


def set_excel_style(filename, cell_s, h_long, d_long, sheet_name='Sheet1'):
    wb = openpyxl.load_workbook(filename)
    sheet = wb[sheet_name]
    # 初始化字体样式
    font_bold = Font(name='Calibri',
                     size=10,
                     bold=True,
                     italic=False,
                     vertAlign=None,
                     underline='none',
                     strike=False,
                     color='00FF0000',
                     outline='None')

    font_not_bold = Font(name='Calibri',
                         size=10,
                         italic=False,
                         vertAlign=None,
                         underline='none',
                         strike=False,
                         color='FF000000',
                         outline='None')

    font_not_bold2 = Font(name='微軟正黑體',
                          size=10,
                          italic=False,
                          vertAlign=None,
                          underline='none',
                          strike=False,
                          color='FF000000',
                          outline='None')

    # 取消网格线
    sheet.views.sheetView[0].showGridLines = False  # 设置不显示网格线

    # 设置默认缩放比例
    sheet.views.sheetView[0].zoomScale = 100  # 设置默认缩放比例

    # 设置边框样式
    # border_thin = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'),
    #                      left=Side(border_style='thin'), right=Side(border_style='thin'))
    border_medium = Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'),
                           left=Side(border_style='medium'), right=Side(border_style='medium'))
    # border_none = Border(top=Side(), bottom=Side(), left=Side(), right=Side())
    # border_double = Border(top=Side(border_style='thin'), bottom=Side(border_style='double'),
    #                        left=Side(), right=Side())

    # 设置整体的字体
    for row in sheet.rows:
        for cell in row:
            cell.font = font_not_bold
            cell.alignment = Alignment(horizontal='center', vertical='center')  # 居中

    # 表头字体红色加粗居中
    for i in sheet.iter_rows(min_row=3, max_row=3):  # 按行列获取数据 row3
        for j in i:
            j.alignment = Alignment(horizontal='center', vertical='center')
            j.font = font_bold
            # j.fill = PatternFill(fill_type='solid', fgColor="FFFF00")
    for i in sheet.iter_rows(min_row=3 + h_long + 5, max_row=3 + h_long + 5):  # 按行列取数据 表头row17
        for j in i:
            j.alignment = Alignment(horizontal='center', vertical='center')
            j.font = font_bold

    # 次表头字体加粗居中
    for i in sheet.iter_rows(min_row=4, max_row=4):  # 按行列获取数据 row4
        for j in i:
            j.font = font_not_bold2
    for i in sheet.iter_rows(min_row=3 + h_long + 6, max_row=3 + h_long + 6):  # 按行列获取数据 row18
        for j in i:
            j.font = font_not_bold2

    # 加粗边框
    for i in sheet.iter_rows(min_row=3, max_row=cell_s + h_long, min_col=1, max_col=4):  # 按行获取数据 A3:D13
        for j in i:
            j.border = border_medium
            j.fill = PatternFill(fill_type='solid', fgColor="F0F8FF")
    for i in sheet.iter_rows(min_row=cell_s + h_long + 4, max_row=cell_s + h_long + d_long + 5, min_col=1,
                             max_col=4):  # 按行获取数据 A18:D28
        for j in i:
            j.border = border_medium
            j.fill = PatternFill(fill_type='solid', fgColor="F0F8FF")
    for i in sheet.iter_rows(min_row=3, max_row=cell_s + h_long, min_col=6, max_col=9):  # 按行获取数据 F3:I13
        for j in i:
            j.border = border_medium
            j.fill = PatternFill(fill_type='solid', fgColor="FFFFE0")
    for i in sheet.iter_rows(min_row=cell_s + h_long + 4, max_row=cell_s + h_long + d_long + 5, min_col=6,
                             max_col=9):  # 按行获取数据 F18:I28
        for j in i:
            j.border = border_medium
            j.fill = PatternFill(fill_type='solid', fgColor="FFFFE0")
    for i in sheet.iter_rows(min_row=3, max_row=cell_s + h_long, min_col=11, max_col=11):  # 按行获取数据 K3:K13
        for j in i:
            j.border = border_medium
            j.fill = PatternFill(fill_type='solid', fgColor="87CEEB")
    for i in sheet.iter_rows(min_row=cell_s + h_long + 4, max_row=cell_s + h_long + d_long + 5, min_col=11,
                             max_col=11):  # 按行获取数据 K18:K28
        for j in i:
            j.border = border_medium
            j.fill = PatternFill(fill_type='solid', fgColor="87CEEB")

    # 行高、列宽
    for i in range(3, sheet.max_row):
        sheet.row_dimensions[i].height = 18
    for i in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(i)].width = 10
    sheet.column_dimensions['E'].width = 6
    sheet.column_dimensions['J'].width = 6
    sheet.column_dimensions['K'].width = 16
    # sheet.row_dimensions[1].height = 30

    # sheet.row_dimensions.height = 40  # 将整个表行高设为50 未执行成功
    # sheet.column_dimensions.width = 30

    # 合并列
    sheet.merge_cells(start_row=5, start_column=3, end_row=cell_s + h_long, end_column=3)
    sheet.merge_cells(start_row=cell_s + h_long + 6, start_column=3, end_row=cell_s + h_long + d_long + 5,
                      end_column=3)
    sheet.merge_cells(start_row=5, start_column=8, end_row=13, end_column=8)
    sheet.merge_cells(start_row=cell_s + h_long + 6, start_column=8, end_row=cell_s + h_long + d_long + 5,
                      end_column=8)
    wb.save(filename)


def write_excel(m_item_dict, f_item_dict, filename, cell_s, h_long10, d_long):
    write_data(m_item_dict, f_item_dict, filename, cell_s, h_long10, d_long)
    set_excel_style(filename, cell_s, h_long10, d_long)


class MainGUI:
    def __init__(self):
        window = Tk()
        window.geometry('600x400')
        window.title('Hinge Plate & Dome shim Data Summary Tool')

        self.label = Label(window, font=("Arial black", 15), fg="black", text='选择源文件目录')
        self.label.place(x=10, y=10)

        # 增加文本框
        self.srcfilepathname = StringVar(value='')  # 文本框默认显示的内容
        self.input_entry = Entry(window,
                                 highlightcolor='red',
                                 highlightthickness=1,
                                 textvariable=self.srcfilepathname)
        self.input_entry.place(x=10, y=40, width=410, height=30)
        self.btn_in = Button(window,
                             text='选择目录',
                             command=self.open_directory,
                             width=10,
                             height=1)
        self.btn_in.place(x=400, y=40)
        self.btn_run = Button(
            window, text="执行并保存", command=self.run, width=10, height=1)
        self.btn_run.place(x=500, y=40)

        # 增加列表框
        self.result_show = Listbox(window, bg='Azure')
        self.result_show.place(x=10, y=100, width=570, height=290)
        self.sbY = Scrollbar(self.result_show,
                             command=self.result_show.yview)  # 在列表框中增加Y轴滚动条
        self.sbX = Scrollbar(self.result_show, command=self.result_show.xview, orient=HORIZONTAL)  # 在列表框中增加X轴滚动条
        self.result_show.config(xscrollcommand=self.sbX.set)
        self.result_show.config(yscrollcommand=self.sbY.set)
        self.sbX.pack(side=BOTTOM, fill=X)
        self.sbY.pack(side=RIGHT, fill=Y)
        window.mainloop()

    def open_directory(self):
        inpath = filedialog.askdirectory(title='选择文件目录')
        self.srcfilepathname.set(inpath)  # 获取变量值传给输入框，显示选择的文件目录，方法一
        # if inpath:
        #     self.input_entry.insert(0, inpath)  # 将选择的文件目录插入输入框，方法二

    def run(self):
        inpath = self.input_entry.get()
        print('path', inpath)
        if not inpath:
            messagebox.showinfo(title='提示', message='文件目录为空，请选择文件目录!')
        else:
            filetypes = [("XLSX", "*.xlsx"), ("PNG", "*.png"), ("JPG", "*.jpg"), ("GIF", "*.gif"),
                         ("txt files", "*.txt"), ('All files', '*')]
            # 返回一个 pathname 文件路径字符串，如果取消或者关闭则返回空字符，返回文件如何操作是后续代码的事情，
            # 该函数知识返回选择文件的文件名字，不具备保存文件的能力
            filenewpathname = filedialog.asksaveasfilename(title='保存文件',
                                                           filetypes=filetypes,
                                                           defaultextension='.xlsx')  # 用户桌面
            # filenewpathname = filedialog.asksaveasfilename(title='保存文件',
            #                                             filetypes=filetypes,
            #                                             defaultextension='.xlsx',
            #                                             initialdir=os.path.expanduser('~') + '\\' + 'Desktop')  # 用户桌面
            self.data_analysis(inpath, filenewpathname)

    #
    # def ck_model(self, filepath):
    #     with open(filepath, 'r') as f:
    #         data = f.read()
    #         m_model = '-M-'
    #         f_model = '-F-'
    #         # if data.find('-M-') >= 0:  # 判断是金属否-M- 方法一
    #         if m_model in data:  # 判断是包含str 方法二
    #             print(f"Model is {m_model}")
    #             return m_model
    #         elif f_model in data:
    #             print(f"Model is {f_model}")
    #             return f_model
    #         else:
    #             print(f"Model is error.")

    # def item_sum(self, value_l, m_str, *parame):
    #     if m_str == 'f':  # fabric小数点后不带0
    #         for j in parame:
    #             for i in j:
    #                 i = m_str + i
    #                 if len(i) >= 7:
    #                     # print(i)
    #                     if i[-1] == '0':
    #                         item_n = i[:len(i) - 1]
    #                         item_v = value_l.count(str(i[3:len(i) - 1]))
    #                         f_item_dict[item_n] = item_v
    #                         f_item_n_list.append(item_n)
    #                         # print(item_dict)
    #                         # print(item_n_list)
    #                         # print(item_dict['fhv0.1'])
    #                     else:
    #                         item_n = i
    #                         item_v = value_l.count(str(i[3:]))
    #                         f_item_dict[item_n] = item_v
    #                         f_item_n_list.append(item_n)
    #
    #                 else:
    #                     if i[-2:] == '.0':
    #                         item_n = i[:-2]
    #                         item_v = value_l.count(str(i[3:-2]))
    #                         f_item_dict[item_n] = item_v
    #                         f_item_n_list.append(item_n)
    #                     else:
    #                         item_n = i
    #                         item_v = value_l.count(str(i[3:]))
    #                         f_item_dict[item_n] = item_v
    #                         f_item_n_list.append(item_n)
    #     elif m_str == 'm':  # metal小数点后带0
    #         for j in parame:
    #             for i in j:
    #                 i = m_str + i
    #                 item_n = i
    #                 item_v = value_l.count(str(i[3:]))
    #                 m_item_dict[item_n] = item_v
    #                 m_item_n_list.append(item_n)
    #     return f_item_n_list, f_item_dict, m_item_n_list, m_item_dict

    # def write_data(self, m_item_dict, f_item_dict, save_filename, cell_s, h_long, d_long):
    #     # cell_s = 4  # 数据行前其它行
    #     # h_long = 9  # HINGE PLATE （metal） data个数
    #     # d_long = 10  # Dome shim（metal） data个数
    #     wb = openpyxl.Workbook()
    #     wb.create_sheet(index=0, title="Sheet1")
    #     sheet = wb['Sheet1']
    #
    #     title = [['HINGE PLATE （metal）', '', '', '', '', 'HINGE PLATE （fabric）', '', '', '', '', 'metal+fabric'],
    #              ['规格', '数量', 'total ', '比例', '', '规格', '数量', 'total ', '比例', '', '总比例']]
    #     # 写入表格栏名
    #     for i in range(0, len(title)):
    #         for j in range(0, len(title[i])):
    #             sheet.cell(row=i + 3, column=j + 1, value=str(title[i][j]))
    #             sheet.cell(row=i + 3 + h_long + 5, column=j + 1, value=str(title[i][j]))
    #
    #     # 合并单元格
    #     sheet.merge_cells('A3:D3')
    #     sheet.merge_cells('F3:I3')
    #     sheet.merge_cells('A{}:D{}'.format(cell_s + h_long + 4, cell_s + h_long + 4))
    #     sheet.merge_cells('F{}:I{}'.format(cell_s + h_long + 4, cell_s + h_long + 4))
    #
    #     '''写入 metal data ROW A-D'''
    #     cell_no = cell_s
    #     for key, value in m_item_dict.items():
    #         cell_no += 1
    #         # HINGE PLATE （metal） data row5~13
    #         if cell_s <= cell_no <= cell_s + h_long:
    #             sheet['A' + str(cell_no)] = key[3:]
    #             sheet['B' + str(cell_no)] = value
    #             sheet['D' + str(cell_no)] = '=B{}/C{}'.format(cell_no, cell_s + 1)  # 比例=B5/C5
    #             sheet['D' + str(cell_no)].number_format = '0.00%'
    #             sheet['K' + str(cell_no)] = '=(B{}+G{})/(C{}+H{})'.format(cell_no, cell_no, cell_s + 1,
    #                                                                       cell_s + 1)  # 总比例=(B5+G5)/(C5+H5)
    #             sheet['K' + str(cell_no)].number_format = '0.00%'
    #         # Dome shim（metal） row19~28
    #         else:
    #             sheet['A' + str(cell_no + 5)] = key[3:]
    #             sheet['B' + str(cell_no + 5)] = value
    #             sheet['D' + str(cell_no + 5)] = '=B{}/C{}'.format(cell_no + 5, cell_s + 1 + h_long + 5)  # 比例=B19/C19
    #             sheet['D' + str(cell_no + 5)].number_format = '0.00%'
    #             # 总比例=(B19+G19)/(C19+H19)
    #             sheet['K' + str(cell_no + 5)] = '=(B{}+G{})/(C{}+H{})'.format(cell_no + 5, cell_no + 5,
    #                                                                           cell_s + 1 + h_long + 5,
    #                                                                           cell_s + 1 + h_long + 5)
    #             sheet['K' + str(cell_no + 5)].number_format = '0.00%'
    #     # sheet['C5'] = '=SUM(B5:B13)'  # 求HINGE 和 方法一
    #     # 求HINGE PLATE （metal）总数和 方法二
    #     sheet['C{}'.format(cell_s + 1)] = '=SUM(B{}:B{})'.format(cell_s + 1, cell_s + h_long)
    #     # 求Dome shim（metal）总数和
    #     sheet['C{}'.format(cell_s + 1 + h_long + 5)] = '=SUM(B{}:B{})'.format(cell_s + 1 + h_long + 5,
    #                                                                           cell_s + h_long + 5 + d_long)
    #
    #     '''写入 fabric data ROW F-G'''
    #     cell_no = cell_s
    #     for key, value in f_item_dict.items():
    #         cell_no += 1
    #         # HINGE PLATE （fabric） data row5~13
    #         if cell_s <= cell_no <= cell_s + h_long:
    #             sheet['F' + str(cell_no)] = key[3:]
    #             sheet['G' + str(cell_no)] = value
    #             sheet['I' + str(cell_no)] = '=G{}/H{}'.format(cell_no, cell_s + 1)  # 比例=G5/H5
    #             sheet['I' + str(cell_no)].number_format = '0.00%'
    #         # Dome shim（fabric） row19~28
    #         else:
    #             sheet['F' + str(cell_no + 5)] = key[3:]
    #             sheet['G' + str(cell_no + 5)] = value
    #             sheet['I' + str(cell_no + 5)] = '=G{}/H{}'.format(cell_no + 5, cell_s + 1 + h_long + 5)  # 比例=G19/H19
    #             sheet['I' + str(cell_no + 5)].number_format = '0.00%'
    #     # 求HINGE PLATE （fabric）总数和
    #     sheet['H{}'.format(cell_s + 1)] = '=SUM(G{}:G{})'.format(cell_s + 1, cell_s + h_long)
    #     # 求Dome shim（fabric）总数和
    #     sheet['H{}'.format(cell_s + 1 + h_long + 5)] = '=SUM(G{}:G{})'.format(cell_s + 1 + h_long + 5,
    #                                                                           cell_s + h_long + 5 + d_long)
    #
    #     wb.save(save_filename)
    #
    # def set_excel_style(self, filename, cell_s, h_long, d_long, sheet_name='Sheet1'):
    #     wb = openpyxl.load_workbook(filename)
    #     sheet = wb[sheet_name]
    #     # 初始化字体样式
    #     font_bold = Font(name='Calibri',
    #                      size=10,
    #                      bold=True,
    #                      italic=False,
    #                      vertAlign=None,
    #                      underline='none',
    #                      strike=False,
    #                      color='00FF0000',
    #                      outline='None')
    #
    #     font_not_bold = Font(name='Calibri',
    #                          size=10,
    #                          italic=False,
    #                          vertAlign=None,
    #                          underline='none',
    #                          strike=False,
    #                          color='FF000000',
    #                          outline='None')
    #
    #     font_not_bold2 = Font(name='微軟正黑體',
    #                           size=10,
    #                           italic=False,
    #                           vertAlign=None,
    #                           underline='none',
    #                           strike=False,
    #                           color='FF000000',
    #                           outline='None')
    #
    #     # 取消网格线
    #     sheet.views.sheetView[0].showGridLines = False  # 设置不显示网格线
    #
    #     # 设置默认缩放比例
    #     sheet.views.sheetView[0].zoomScale = 100  # 设置默认缩放比例
    #
    #     # 设置边框样式
    #     # border_thin = Border(top=Side(border_style='thin'), bottom=Side(border_style='thin'),
    #     #                      left=Side(border_style='thin'), right=Side(border_style='thin'))
    #     border_medium = Border(top=Side(border_style='medium'), bottom=Side(border_style='medium'),
    #                            left=Side(border_style='medium'), right=Side(border_style='medium'))
    #     # border_none = Border(top=Side(), bottom=Side(), left=Side(), right=Side())
    #     # border_double = Border(top=Side(border_style='thin'), bottom=Side(border_style='double'),
    #     #                        left=Side(), right=Side())
    #
    #     # 设置整体的字体
    #     for row in sheet.rows:
    #         for cell in row:
    #             cell.font = font_not_bold
    #             cell.alignment = Alignment(horizontal='center', vertical='center')  # 居中
    #
    #     # 表头字体红色加粗居中
    #     for i in sheet.iter_rows(min_row=3, max_row=3):  # 按行列获取数据 row3
    #         for j in i:
    #             j.alignment = Alignment(horizontal='center', vertical='center')
    #             j.font = font_bold
    #             # j.fill = PatternFill(fill_type='solid', fgColor="FFFF00")
    #     for i in sheet.iter_rows(min_row=3 + h_long + 5, max_row=3 + h_long + 5):  # 按行列取数据 表头row17
    #         for j in i:
    #             j.alignment = Alignment(horizontal='center', vertical='center')
    #             j.font = font_bold
    #
    #     # 次表头字体加粗居中
    #     for i in sheet.iter_rows(min_row=4, max_row=4):  # 按行列获取数据 row4
    #         for j in i:
    #             j.font = font_not_bold2
    #     for i in sheet.iter_rows(min_row=3 + h_long + 6, max_row=3 + h_long + 6):  # 按行列获取数据 row18
    #         for j in i:
    #             j.font = font_not_bold2
    #
    #     # 加粗边框
    #     for i in sheet.iter_rows(min_row=3, max_row=cell_s + h_long, min_col=1, max_col=4):  # 按行获取数据 A3:D13
    #         for j in i:
    #             j.border = border_medium
    #             j.fill = PatternFill(fill_type='solid', fgColor="F0F8FF")
    #     for i in sheet.iter_rows(min_row=cell_s + h_long + 4, max_row=cell_s + h_long + d_long + 5, min_col=1,
    #                              max_col=4):  # 按行获取数据 A18:D28
    #         for j in i:
    #             j.border = border_medium
    #             j.fill = PatternFill(fill_type='solid', fgColor="F0F8FF")
    #     for i in sheet.iter_rows(min_row=3, max_row=cell_s + h_long, min_col=6, max_col=9):  # 按行获取数据 F3:I13
    #         for j in i:
    #             j.border = border_medium
    #             j.fill = PatternFill(fill_type='solid', fgColor="FFFFE0")
    #     for i in sheet.iter_rows(min_row=cell_s + h_long + 4, max_row=cell_s + h_long + d_long + 5, min_col=6,
    #                              max_col=9):  # 按行获取数据 F18:I28
    #         for j in i:
    #             j.border = border_medium
    #             j.fill = PatternFill(fill_type='solid', fgColor="FFFFE0")
    #     for i in sheet.iter_rows(min_row=3, max_row=cell_s + h_long, min_col=11, max_col=11):  # 按行获取数据 K3:K13
    #         for j in i:
    #             j.border = border_medium
    #             j.fill = PatternFill(fill_type='solid', fgColor="87CEEB")
    #     for i in sheet.iter_rows(min_row=cell_s + h_long + 4, max_row=cell_s + h_long + d_long + 5, min_col=11,
    #                              max_col=11):  # 按行获取数据 K18:K28
    #         for j in i:
    #             j.border = border_medium
    #             j.fill = PatternFill(fill_type='solid', fgColor="87CEEB")
    #
    #     # 行高、列宽
    #     for i in range(3, sheet.max_row):
    #         sheet.row_dimensions[i].height = 18
    #     for i in range(1, sheet.max_column + 1):
    #         sheet.column_dimensions[get_column_letter(i)].width = 10
    #     sheet.column_dimensions['E'].width = 6
    #     sheet.column_dimensions['J'].width = 6
    #     sheet.column_dimensions['K'].width = 16
    #     # sheet.row_dimensions[1].height = 30
    #
    #     # sheet.row_dimensions.height = 40  # 将整个表行高设为50 未执行成功
    #     # sheet.column_dimensions.width = 30
    #
    #     # 合并列
    #     sheet.merge_cells(start_row=5, start_column=3, end_row=cell_s + h_long, end_column=3)
    #     sheet.merge_cells(start_row=cell_s + h_long + 6, start_column=3, end_row=cell_s + h_long + d_long + 5,
    #                       end_column=3)
    #     sheet.merge_cells(start_row=5, start_column=8, end_row=13, end_column=8)
    #     sheet.merge_cells(start_row=cell_s + h_long + 6, start_column=8, end_row=cell_s + h_long + d_long + 5,
    #                       end_column=8)
    #     wb.save(filename)

    # def write_excel(self, m_item_dict, f_item_dict, filename, cell_s, h_long10, d_long):
    #     write_data(m_item_dict, f_item_dict, filename, cell_s, h_long10, d_long)
    #     set_excel_style(filename, cell_s, h_long10, d_long)

    def data_analysis(self, inpath, filename):
        m_value_l = []
        f_value_l = []
        in_path = Path(inpath)
        exts = ['.txt']
        # print(in_path)
        # out_path = in_path.with_name(in_path.stem + '-resize')  # 新建加-resize目录 方法一
        # # out_path = in_path.parent / (in_path.stem+'-resize')  # 新建加-resize目录 方法二
        # if not out_path.exists():
        #     out_path.mkdir()
        self.result_show.delete(0, 'end')  # 删除所有元素
        for root, dirs, files in walk(in_path):
            # inroot_path = Path(root)
            for file in files:
                file_rel_path = Path(root).joinpath(file)
                ext2 = file_rel_path.suffix
                ext = ext2.lower()
                if ext not in exts:  # 不是txt文件不处理
                    continue
                print('---------')
                print(file_rel_path)
                self.result_show.insert("end", file_rel_path)
                model = ck_model(file_rel_path)
                if model == '-M-':
                    with open(file_rel_path, "r") as f:
                        data = f.readlines()  # 每行做为列表元素
                        for word_s in data:
                            word_l = word_s.split('\t')
                            m_value_l.extend([word_l[5], word_l[6], word_l[7]])
                            # print(m_value_l)
                if model == '-F-':
                    with open(file_rel_path, "r") as f:
                        data = f.readlines()  # 每行做为列表元素
                        for word_s in data:
                            word_l = word_s.split('\t')
                            f_value_l.extend([word_l[5], word_l[6], word_l[7]])
                            # print(m_value_l)
                # m_value_item = ['mhv0', 'mhv005', 'mhv010']
            item_sum(f_value_l, 'f', source_item_n)
            print('Fabric item name :', f_item_n_list_)
            print('Fabric item data:', f_item_dict_)
            item_sum(m_value_l, 'm', source_item_n)
            print('Metal item name :', m_item_n_list_)
            print('Metal item data:', m_item_dict_)
            write_excel(m_item_dict_, f_item_dict_, filename, cell_s_, h_long_, d_long_)
        messagebox.showinfo(title='提示', message='处理完成!')


if __name__ == '__main__':
    gui = MainGUI()
