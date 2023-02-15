# _*_ coding: utf-8 _*_
# @Time   : 2022/6/24 下午 03:09
# @Author : 未来战士yuyu!!
# @FileName : 節能.py
# @Software : PyCharm
# @Blog     : http://blog.csdn.net/u010105243/article/
# Change List:
# 061002 first release.
# 062001 tkinter導致RuntimeError: main thread is not in main，加root.destroy()
# 062201 添加download刷卡数据, 加防锁屏
# 062301 修改edgedriver路径
# 062401 導入制定庫, 加wait_file_until
# 062501 加wait_tmp_until 取消copy_serial_fail,修改merged_file直接到eAttance,刷卡加可选部门,刷卡资料两天一起拉
# 062601 刷卡加可选部门,取消刷卡资料两天一起拉，加責任制
# 062801 for dx
# 070701 dlsk延时8s
# 070801打开网页后添加WebDriverWait(driver, 20).until(EC.visibility_of_element_located(loc))
# 071201 unlock_scr 加pyautogui.FAILSAFE = False防止鼠标放到角落会报错退出,加部门输入，加支援部剔除長期支援
# 072801 修改弹窗警示方式自动关闭弹窗，加用户名密码自动登入网页
# 072801 取消加用户名密码自动登入网页
# 072901 完成后加提示画面
# 080101 有刷卡有点名时取消比对休假，否则周一如有夜班转白班会报有刷卡休假错误
# 080801 多线程报错main thread is not in main loop,删除pysimplegui警示窗口
# 080901  job_thread.setDaemon(True) # 创建多线程 设置以保护模式启动，即主线程运行结束，子线程也停止运行,today_str加入主线程里
# 082001 剔除部分责任制人员不统计如厂长，设为工作日执行
# 082201 增加责任制移除人员
# 082601 wait_file_until检查下载文件名，取消检查未下载文件后缀(*.tmp,*.crdownload)
# 083001 下载等待时间加长
# 091601 改用Chrome浏览器
# 092301 删除出勤异常数据重复人员
# 101701 查找时将工程改为大写
# 110101 將'責任制名單.xlsx'文件改為'全部責任制名單.xlsx'和'需統計出勤責任制名單.xlsx'，方便不統計某位責任制人員，只需修改文件不用修改程式
# 112901 增加指定发件人
# 2022121606 检查设定的发件人与outlook读出来的是否一致

from win32api import SetCursorPos, mouse_event
from win32gui import FindWindow, IsWindow, IsWindowEnabled, IsWindowVisible, ShowWindow, \
    SetWindowPos, GetWindowRect
from win32con import SW_SHOWNORMAL, HWND_TOPMOST, SWP_NOMOVE, SWP_NOACTIVATE, SWP_NOOWNERZORDER, \
    SWP_SHOWWINDOW, SWP_NOSIZE, MOUSEEVENTF_LEFTDOWN, MOUSEEVENTF_LEFTUP
from time import sleep
from openpyxl import load_workbook, Workbook
from selenium import webdriver
from selenium.webdriver.edge.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.select import Select
from datetime import date, datetime, timedelta
import os
from pandas import read_excel, DataFrame, concat, ExcelFile
from shutil import copyfile
from win32com.client import Dispatch
import pythoncom
from re import fullmatch, compile
from codecs import open
import glob
from schedule import every, run_pending
from threading import Thread
from numpy import where
import pyautogui
from PySimpleGUI import popup_auto_close

user_name = os.path.expanduser('~')
dl_di = user_name + '\\' + 'Downloads'
file_folder = dl_di + '\\' + 'eAttendance'

# dept_dianm = ['B5E3400M0E', 'B5E3750M0E']  # 'B5E3400M0E',工程 'B5E3750M0E' 维修组
# dept_renli = ['0', '1', '5']   # 0支援部-製三課(工程組)，1 工程剖, 5 專案規劃部-維修課

# 三厂 for apple
dept_dianm = ['(名碩)第五事業處-製造二處-支援部-製三課', '(名碩)第五事業處-製造三處-三廠']
dept_renli = ['(名碩)第五事業處-製造二處-支援部-製三課', '(名碩)第五事業處-製造三處-三廠']
dept_shuaka = ['ALL']

# # 三厂 for Jing6_Ye
# dept_dianm = ['(名碩)第五事業處-製造二處-支援部-製二課', '(名碩)第五事業處-製造三處-二廠']
# dept_renli = ['(名碩)第五事業處-製造二處-支援部-製二課', '(名碩)第五事業處-製造三處-二廠']
# dept_shuaka = ['ALL']

# # 三厂-工程 for dongxia
# dept_dianm = ['(名碩)第五事業處-製造三處-三廠-工程部', '(名碩)第五事業處-製造三處-三廠-專案規劃部-維修課']
# dept_renli = ['(名碩)第五事業處-製造三處-三廠-工程部', '(名碩)第五事業處-製造二處-支援部-製三課(工程組)', '(名碩)第五事業處-製造三處-三廠-專案規劃部-維修課']
# dept_shuaka = ['B5E3410M0E_(名碩)第五事業處-製造三處-三廠-工程部-工程一課', 'B5E3420M0E_(名碩)第五事業處-製造三處-三廠-工程部-工程二課',
#                'B5E3430M0E_(名碩)第五事業處-製造三處-三廠-工程部-FA課', 'B5E3750M0E_(名碩)第五事業處-製造三處-三廠-專案規劃部-維修課',
#                'B5B0B3CM0E_(名碩)第五事業處-製造二處-支援部-製三課(工程組)']

EX_MAILTO = ''  # 收件者
EX_MAILCC = ''  # 副本
MAIL_PATH = file_folder + '\\' + 'MailAddressListTemplet.xlsx'  # 邮件地址文件路徑
# edge_driver = file_folder + '\\' + 'msedgedriver.exe'
browser_driver = file_folder + '\\' + 'chromedriver.exe'


def dept_input(dept_list, dept_str, exp):
    print(dept_str + ' 的部门名称为: %s：' % dept_list)
    print('重输请输入r,确认请输入q')
    ipt2 = input()
    if ipt2 == 'r':
        dept_list = []
        while True:
            print('请依次分别输入欲查询 ' + dept_str + ' 的部门名称，如 ' + exp + '，结束输入请输入q：')
            ipt = input()
            if ipt != 'q':
                dept_list.append(ipt)
            else:
                print(dept_str + '的部门名称为: %s：' % dept_list)
                print('重输请输入r,确认请输入q')
                ipt2 = input()
                if ipt2 == 'r':
                    dept_list = []
                elif ipt2 == 'q':
                    return dept_list
    else:
        return dept_list


def getdate(before_n_day):
    today_tm = datetime.now()
    offset = timedelta(days=-before_n_day)  # 计算偏移量
    re_date = (today_tm + offset).strftime('%Y/%m/%d')  # 获取想要的日期的时间
    return re_date


def msg_alert_psgui():
    popup_auto_close("Python自动流程即将开始，请勿动鼠标与键盘！~")
    popup_auto_close("Python自动流程即将开始，请勿动鼠标与键盘！~")
    popup_auto_close("Python自动流程即将开始，请勿动鼠标与键盘！~")
    popup_auto_close("Python自动流程即将开始，请勿动鼠标与键盘！~")
    popup_auto_close("Python自动流程即将开始，请勿动鼠标与键盘！~")


def complete_alert_psgui():
    popup_auto_close("Python 出勤统计完成！~")
    popup_auto_close("Python 出勤统计完成！~")
    popup_auto_close("Python 出勤统计完成！~")
    popup_auto_close("Python 出勤统计完成！~")
    popup_auto_close("Python 出勤统计完成！~")


def alert_cfm_win(find_window):
    while True:
        try:
            sleep(10)
            # 置顶窗口
            # print("置顶窗口")
            # 窗口需要正常大小且在后台，不能最小化
            hwnd = FindWindow(None, find_window)
            if (IsWindow(hwnd)
                    and IsWindowEnabled(hwnd)
                    and IsWindowVisible(hwnd)):

                # 激活显示窗口，使其成为置顶活动窗口
                ShowWindow(hwnd, SW_SHOWNORMAL)
                # SetForegroundWindow(hwnd)
                # 置顶
                SetWindowPos(hwnd, HWND_TOPMOST, 0, 0, 0, 0,
                             SWP_NOMOVE | SWP_NOACTIVATE | SWP_NOOWNERZORDER | SWP_SHOWWINDOW | SWP_NOSIZE)
                # 取消置顶
                # SetWindowPos(hwnd, HWND_NOTOPMOST, 0, 0, 0, 0,SWP_SHOWWINDOW|SWP_NOSIZE|SWP_NOMOVE)

                # 获取窗口的位置信息
                left, top, right, bottom = GetWindowRect(hwnd)
                print(left, top, right, bottom)
                x = int(left + (right - left) // 1.25)
                y = int(top + (bottom - top) // 6 * 5.2)
                print(x, y)
                sleep(1)
                SetCursorPos([x, y])
                sleep(0.1)
                sleep(5)
                mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
                sleep(0.1)
                mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
                sleep(0.3)
                print('Finded %s window and clicked!' % find_window)
            else:
                pass
        except Exception as alert_cfm_win_er:
            print('alert_cfm_win error:%s' % alert_cfm_win_er)


def mail_cfm_win(find_window):
    while True:
        try:
            sleep(10)
            # 置顶窗口
            # print("置顶窗口")
            # 窗口需要正常大小且在后台，不能最小化
            hwnd = FindWindow(None, find_window)
            if (IsWindow(hwnd)
                    and IsWindowEnabled(hwnd)
                    and IsWindowVisible(hwnd)):

                # 激活显示窗口，使其成为置顶活动窗口
                ShowWindow(hwnd, SW_SHOWNORMAL)
                # SetForegroundWindow(hwnd)
                # 置顶
                SetWindowPos(hwnd, HWND_TOPMOST, 0, 0, 0, 0,
                             SWP_NOMOVE | SWP_NOACTIVATE | SWP_NOOWNERZORDER | SWP_SHOWWINDOW | SWP_NOSIZE)
                # 取消置顶
                # SetWindowPos(hwnd, HWND_NOTOPMOST, 0, 0, 0, 0,SWP_SHOWWINDOW|SWP_NOSIZE|SWP_NOMOVE)

                # 获取窗口的位置信息
                left, top, right, bottom = GetWindowRect(hwnd)
                print(left, top, right, bottom)
                x = int(left + (right - left) // 3.7)
                y = int(top + (bottom - top) // 6 * 5.2)
                print(x, y)
                sleep(1)
                SetCursorPos([x, y])
                sleep(0.1)
                mouse_event(MOUSEEVENTF_LEFTDOWN, 0, 0, 0, 0)
                sleep(0.1)
                mouse_event(MOUSEEVENTF_LEFTUP, 0, 0, 0, 0)
                sleep(0.3)
                print('Finded %s window and clicked!' % find_window)
            else:
                pass
        except Exception as mail_cfm_win_er:
            print('mail_cfm_win error: %s' % mail_cfm_win_er)


def if_exist_foler(file_path):  # 判断是否存在文件夹
    if os.path.exists(file_path):  # 该文件路径存在"
        if os.path.isdir(file_path):  # 判断是文件夹
            pass
        else:
            os.mkdir(file_path)
    else:
        os.mkdir(file_path)  # 路径不存在，创建文件目录


def del_serial_file(path, com_str, ext_name):
    all_name = glob.glob(path + '\\' + '*' + com_str + '*.' + ext_name)
    for name in all_name:
        base_name = os.path.basename(name)
        os.remove(path + '\\' + base_name)


def wait_tmp_until(folder, com_str, ext_nm):
    all_name = glob.glob(folder + '\\' + '*' + com_str + '*.' + ext_nm)
    print('Waiting for file to be downloaded completely.')
    for name in all_name:
        base_name = os.path.basename(name)
        file_nm = (folder + '\\' + base_name)
        while True:
            sleep(5)
            if os.path.exists(file_nm):
                print('Waiting for file to be downloaded completely.')
            else:
                print('File has been downloaded completely.')
                break


def wait_file_until(foleder, file_nm, ext_nm, list_ele, l_ele_idx):
    print(list_ele, l_ele_idx)
    while True:
        sleep(3)
        if l_ele_idx == 0:
            print('Waiting for %s to be downloaded completely.' % (
                    foleder + '\\' + file_nm + '.' + ext_nm))
            if os.path.exists(foleder + '\\' + file_nm + '.' + ext_nm):
                print('%s has been downloaded.' % (foleder + '\\' + file_nm + '.' + ext_nm))
                break
        else:
            print('Waiting for %s to be downloaded completely.' % (
                    foleder + '\\' + file_nm + ' (' + str(l_ele_idx) + ').' + ext_nm))
            if os.path.exists(foleder + '\\' + file_nm + ' (' + str(l_ele_idx) + ').' + ext_nm):
                print('%s has been downloaded.' % (foleder + '\\' + file_nm + ' (' + str(l_ele_idx) + ').' + ext_nm))
                break


def dl_data():
    del_serial_file(dl_di, 'Attendance', 'xlsx')
    del_serial_file(dl_di, 'Data', 'xlsx')
    del_serial_file(dl_di, '', 'tmp')
    today = date.today()
    today_str = today.strftime('%Y%m%d')
    url = 'http://eattendance.sz.pegatroncorp.com/eAttendance/Index.aspx'

    # # 用户名密码自动登入网页，（关闭proxy或需要在本机没有自动登入的情况才可使用）
    # url = 'http://{}:{}@{}'.format('Apple_Zhang@intra.pegatroncorp.com', '*963.*963.*963.a',
    #                                'eattendance.sz.pegatroncorp.com/eAttendance/Index.aspx')

    # edge浏览器
    # svs = Service(edge_driver)
    # driver = webdriver.Edge(service=svs)  # 打包換電腦后將msedgedriver.exe複製到相應路徑能正常測試
    # # driver = webdriver.Edge()    # 打包換電腦會顯示webdirver PATH錯誤

    # Chrome浏览器
    svs = Service(browser_driver)
    driver = webdriver.Chrome(service=svs)

    driver.maximize_window()  # 浏览器最大化
    driver.implicitly_wait(10)  # 隐性等待10s
    driver.get(url)
    sleep(3)
    # 人力清册
    print(driver.current_window_handle)
    ele = (By.XPATH, '//*[@id="Menu1n0"]/table/tbody/tr/td[1]/a')
    WebDriverWait(driver, 20).until(EC.visibility_of_element_located(ele))
    shezhi = driver.find_element(By.XPATH, '//*[@id="Menu1n0"]/table/tbody/tr/td[1]/a')  # 查看人力清册
    ActionChains(driver).move_to_element(shezhi).perform()  # 悬浮下拉列表
    loc = (By.XPATH, '//*[@id="Menu1n9"]/td/table/tbody/tr/td/a')
    WebDriverWait(driver, 20).until(EC.visibility_of_element_located(loc))
    driver.find_element(By.XPATH, '//*[@id="Menu1n9"]/td/table/tbody/tr/td/a').click()  # 懸浮下拉列表-人力清册
    sleep(3)
    # 要定位一个元素时，怎么都定位不到的时候就要考虑是不是浏览器内嵌了一个 frame 窗口或者要找的元素在新打开的窗口里。这时候就需要进行 frame
    # 的切换或者窗口的切换。
    # iframe的多种切换方式
    # driver.switch_to.frame(0)  # index：传入整型的参数，从 0 开始，这里的 0 就是第一个 frame
    # driver.switch_to.frame("c_Content")    # id：iframe 的 id
    driver.switch_to.frame("c_Content")  # name: iframe 的 name
    # driver.switch_to.frame(driver.find_element(By.TAG_NAME, 'iframe'))  # WebElement: 传入 selenium.webelement 对象
    """
    下拉列表，select元素，处理select元素
    """
    # select 下拉列表方法一
    # bumen_loc = (By.XPATH, '//select[@name="DeptCodeDropDownList"]') # select 下拉列表方法一
    # WebDriverWait(driver, 30).until(EC.visibility_of_element_located(bumen_loc))
    # bumen = driver.find_element(*bumen_loc)
    # 下拉列表方法二
    for i in dept_renli:
        rl_idx = dept_renli.index(i)
        bumen = driver.find_element(By.XPATH, '//*[@id="DeptIdDropDownList"]')  # select 下拉列表方法二 部門
        s = Select(bumen)
        # s.select_by_value(i)  # 部门  by value
        s.select_by_visible_text(i)  # by text  # 部門
        # s.select_by_index(i)     # by index 第一条输入0
        sleep(1)
        chaxun = driver.find_element(By.XPATH, '//*[@id="QueryButton"]')  # 查詢
        chaxun.click()
        sleep(8)
        huichu = driver.find_element(By.XPATH, '//*[@id="ExportButton"]')  # 匯出Excel
        huichu.click()
        sleep(10)
        wait_file_until(dl_di, 'Data', 'xlsx', i, rl_idx)
        # wait_tmp_until(dl_di, '', 'tmp')
        # wait_tmp_until(dl_di, '', 'crdownload')

    # 点名回报
    handles = driver.window_handles
    print(driver.current_window_handle)
    driver.switch_to.window(handles[-1])  # 切回原主页
    shezhi = driver.find_element(By.XPATH, '//*[@id="Menu1n2"]/table/tbody/tr/td[1]/a')  # 点名回报
    ActionChains(driver).move_to_element(shezhi).perform()  # 悬浮下拉列表
    loc = (By.XPATH, '//*[@id="Menu1n15"]/td/table/tbody/tr/td/a')
    WebDriverWait(driver, 20).until(EC.visibility_of_element_located(loc))
    driver.find_element(By.XPATH, '//*[@id="Menu1n15"]/td/table/tbody/tr/td/a').click()  # 懸浮下拉列表-异常处理
    sleep(3)
    driver.switch_to.frame("c_Content")  # name: iframe 的 name
    # 下拉列表方法二
    for i in dept_dianm:
        dm_idx = dept_dianm.index(i)
        bumen = driver.find_element(By.XPATH, '//*[@id="DeptCodeDropDownList"]')  # select 下拉列表方法二 部門
        s = Select(bumen)
        # s.select_by_value(i)  # 部门  by value
        s.select_by_visible_text(i)  # by text  # 部門
        # s.select_by_index(i)     # by index 第一条输入0
        sleep(1)
        driver.find_element(By.XPATH, '//*[@id="ServerityRadioButtonList_2"]').click()  # 全部點名結果
        driver.find_element(By.XPATH, '//*[@id="QueryButton"]').click()  # 查詢
        sleep(8)
        driver.find_element(By.XPATH, '//*[@id="ExportButton"]').click()  # 匯出Excel
        sleep(10)
        wait_file_until(dl_di, today_str + 'Attendance', 'xlsx', i, dm_idx)
        # wait_tmp_until(dl_di, '', 'tmp')  # 未下载完成文件
        # wait_tmp_until(dl_di, '', 'crdownload')  # 未下载完成文件
    driver.close()


def dl_sk_data():
    yesterday_str = getdate(1)
    today = date.today()
    today_ri = today.strftime('%Y/%m/%d')
    riqi_list = [today_ri, yesterday_str]
    del_serial_file(dl_di, 'Result', 'xls')
    del_serial_file(dl_di, '', 'tmp')
    url = 'http://eweb.sz.pegatroncorp.com/ePSZ/CLSS.aspx'

    # # 用户名密码自动登入网页（关闭proxy或需要在本机没有自动登入的情况才可使用）
    # url = 'http://{}:{}@{}'.format('apple_zhang', '*963.*963.*963.a',
    #                                'eweb.sz.pegatroncorp.com/ePSZ/CLSS.aspx')

    # edge浏览器
    # svs = Service(edge_driver)
    # driver = webdriver.Edge(service=svs)  # 打包換電腦后將msedgedriver.exe複製到相應路徑能正常測試
    # # driver = webdriver.Edge()    # 打包換電腦會顯示webdirver PATH錯誤

    # Chrome浏览器
    svs = Service(browser_driver)
    driver = webdriver.Chrome(service=svs)

    driver.maximize_window()  # 浏览器最大化
    driver.implicitly_wait(10)  # 隐性等待10s
    driver.get(url)
    sleep(3)
    # 刷卡
    print(driver.current_window_handle)
    ele = (By.XPATH, '//*[@id="titlepath"]/ul/li')
    WebDriverWait(driver, 20).until(EC.visibility_of_element_located(ele))  # 等待页面加载完成
    # driver.find_element(By.XPATH, '//*[@id="titlepath"]/ul/li').click()
    driver.switch_to.frame("ifrmmain")
    driver.find_element(By.XPATH, '//*[@id="form1"]/ul/li[1]').click()  # 非責任制出勤(主管/人事/助理)
    sleep(1)
    driver.switch_to.default_content()  # 先跳回到最外层的页面
    driver.switch_to.frame("ifrmmain")  # 再切回嵌套的frame
    driver.find_element(By.XPATH, '//*[@id="form1"]/ul/li[1]').click()  # 出勤資料
    sleep(1)
    driver.switch_to.default_content()  # 先跳回到最外层的页面
    driver.switch_to.frame("ifrmmain")  # 再切回嵌套的frame
    driver.find_element(By.XPATH, '//*[@id="form1"]/ul/li[1]').click()  # 非責任制出勤記錄查詢
    sleep(1)
    driver.switch_to.default_content()  # 先跳回到最外层的页面
    driver.switch_to.frame("ifrmmain")  # 再切回嵌套的frame
    for i in dept_shuaka:
        bumen = driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_ddlDeptID"]')  # select 部门
        s = Select(bumen)
        # s.select_by_value(i)  # 部门  by value
        s.select_by_visible_text(i)  # by text  # 部門
        # s.select_by_index(i)     # by index 第一条输入0
        sleep(1)
        # 昨日今日数据
        for d in riqi_list:
            riqi_idx = riqi_list.index(d)
            driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtSDate"]').clear()  # 清空开始日期
            sleep(1)
            driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtSDate"]').send_keys(d)  # 开始日期
            sleep(1)
            driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtEDate"]').clear()  # 清空结束日期
            sleep(1)
            driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_txtEDate"]').send_keys(d)  # 结束日期
            sleep(1)
            driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnQuery"]').click()  # 查询
            sleep(8)
            driver.find_element(By.XPATH, '//*[@id="ctl00_ContentPlaceHolder1_btnExcel"]').click()  # 匯出Excel
            sleep(10)
            wait_file_until(dl_di, 'Result', 'xls', d, riqi_idx)
            # wait_tmp_until(dl_di, '', 'tmp')  # 未下载完成文件
            # wait_tmp_until(dl_di, '', 'crdownload')  # 未下载完成文件
    driver.close()


def copy_serial_file(path, com_str, ext_name):
    del_serial_file(path + '\\' + 'eAttendance', com_str, ext_name)
    all_name = glob.glob(path + '\\' + '*' + com_str + '*.' + ext_name)
    for name in all_name:
        base_name = os.path.basename(name)
        copyfile(path + '\\' + base_name, path + '\\' + 'eAttendance' + '\\' + base_name)


def merge_excel(src_path, dst_path, com_str, ext_name, merged_f):
    if os.path.exists(dst_path + '\\' + merged_f):
        os.remove(dst_path + '\\' + merged_f)
    all_name = glob.glob(src_path + '\\' + '*' + com_str + '*.' + ext_name)
    dfs = []  # 新建列表存放每个文件数据(依次读取多个相同结构的Excel文件并创建DataFrame)
    for name in all_name:
        base_name = os.path.basename(name)
        print(base_name)
        # copyfile(path + '\\' + base_name, path + '\\' + 'Attendance' + '\\' + base_name)
        df = read_excel(src_path + '\\' + base_name)  # 将excel转换成DataFrame
        dfs.append(df)
    try:
        df = concat(dfs)  # 将多个DataFrame合并为一个
        df.to_excel(dst_path + '\\' + merged_f, index=False)  # 写入excel文件，不包含索引数据
    except Exception as merge_er:
        print('merge_excel error: %s' % merge_er)
        sleep(10)
        merge_excel(src_path, dst_path, com_str, ext_name, merged_f)


def dianm_shaix(src_path, dst_path, merged_f):
    if os.path.exists(dst_path + '\\' + merged_f):
        os.remove(dst_path + '\\' + merged_f)
    merge_excel(src_path, dst_path, 'Attendance', 'xlsx', 'merged_Attendance.xlsx')  # 汇整点名数据
    df = DataFrame(read_excel(dst_path + '\\' + 'merged_Attendance.xlsx'))
    # 剔除長期支援
    df = df.loc[~(df['異常原因'] == '長期支援')]
    # 移除厂长、stanly、Alan_Tai、lucy_yan、(注意工号大写，需与点名系统一致）
    df = df.loc[~(df['工號'].isin(['LA0800441', 'LA0801190', 'LA1101483', 'S09198342']))]
    df.to_excel(dst_path + '\\' + merged_f, index=False)  # 写入excel文件，不包含索引数据


def shuaka_shaix(src_path, dst_path, merged_f):
    today = date.today()
    y_day = getdate(1)
    to_day = today.strftime('%Y/%m/%d')
    if os.path.exists(dst_path + '\\' + merged_f):
        os.remove(dst_path + '\\' + merged_f)
    # df = DataFrame(read_excel(day_f, sheet_name='Sheet1'))
    merge_excel(src_path, dst_path, 'Result', 'xls', 'merged_Result.xlsx')
    df = DataFrame(read_excel(dst_path + '\\' + 'merged_Result.xlsx'))
    # 删除包含空格的行，方法一
    # df = df['上班時間'].astype(bool)
    # df =df[df['上班時間'].astype(bool)]
    # 删除包含空格的行，方法二
    df = df.dropna(axis=0, subset=["上班時間"])
    df['出勤卡機狀態'] = '有刷卡'
    df['班別名稱S'] = where(df['班別名稱'].isin(['夜班四', '夜班六', '夜班八']), '夜班', '白班')
    # 白班
    df_d = df.loc[
        (df['班別名稱S'] == '白班') & (df['日期'] == to_day), ['日期', '部門代碼', '部門', '工號', '姓名', '上班時間', '班別名稱', '班別名稱S',
                                                       '出勤卡機狀態']]
    df_d.to_excel(dst_path + '\\' + 'result_sx_sk_day.xlsx', index=False)  # 写入excel文件，不包含索引数据
    # 夜班
    df_n = df.loc[(df['班別名稱S'] == '夜班') & (df['日期'] == y_day), ['日期', '部門代碼', '部門', '工號', '姓名', '上班時間', '班別名稱', '班別名稱S',
                                                                '出勤卡機狀態']]
    df_n.to_excel(dst_path + '\\' + 'result_sx_sk_night.xlsx', index=False)  # 写入excel文件，不包含索引数据
    # 彙總白夜班資料
    merge_excel(dst_path, dst_path, 'result_sx_sk', 'xlsx', merged_f)
    # del_serial_file(dst_path, 'result_sx_sk', 'xlsx')
    # del_serial_file(dst_path, 'merged_Result', 'xlsx')


# 责任制未点名人员
def compare_zrzrl_dm(src_file, dl_file, path):
    today = date.today()
    today_str = today.strftime('%Y%m%d')
    if os.path.exists(path + '\\' + today_str + '未點名人員名單-責任制.xlsx'):
        os.remove(path + '\\' + today_str + '未點名人員名單-責任制.xlsx')
    wb1 = load_workbook(src_file)
    sheet1 = wb1.active
    wb2 = load_workbook(dl_file)
    sheet2 = wb2.active
    diff = [['工号', '姓名', '部门代码', '部门名称']]
    list1 = []
    for i in range(1, sheet1.max_row):
        src = sheet1.cell(row=i + 1, column=1).value.capitalize()  # (capitalize工号首字母大写）
        list1.append(src)
        src = sheet1.cell(row=i + 1, column=2).value
        list1.append(src)
        src = sheet1.cell(row=i + 1, column=3).value
        list1.append(src)
        src = sheet1.cell(row=i + 1, column=4).value
        list1.append(src)
    print('srczrzrl_list' + str(list1))
    list2 = []
    for i in range(1, sheet2.max_row):
        chk = sheet2.cell(row=i + 1, column=12).value
        list2.append(chk)
        chk = sheet2.cell(row=i + 1, column=13).value
        list2.append(chk)
        chk = sheet2.cell(row=i + 1, column=8).value.capitalize()  # (capitalize工号首字母大写）
        list2.append(chk)
        chk = sheet2.cell(row=i + 1, column=9).value
        list2.append(chk)
    print('dlddm_list' + str(list2))
    for i in range(0, len(list1), 4):
        if list1[i] not in list2:
            diff_s = []
            diff_s.append(list1[i])
            diff_s.append(list1[i + 1])
            diff_s.append(list1[i + 2])
            diff_s.append(list1[i + 3])
            diff.append(diff_s)
    print('zrz_diff')
    print(diff)
    #   将数据写入excel
    wb3 = Workbook()
    sheet = wb3.active
    sheet.title = '未點名人員名單-責任制'
    for i in diff:
        sheet.append(i)
    wb3.save(path + '\\' + today_str + '未點名人員名單-責任制.xlsx')


# 非责任制未刷卡点名
def comp_sk_dm(src_file, dl_file, path):
    today = date.today()
    today_str = today.strftime('%Y%m%d')
    if os.path.exists(path + '\\' + today_str + '刷卡點名異常明細-All.xlsx'):
        os.remove(path + '\\' + today_str + '刷卡點名異常明細-All.xlsx')
    if os.path.exists(path + '\\' + today_str + '人力-非責任制.xlsx'):
        os.remove(path + '\\' + today_str + '人力-非責任制.xlsx')
    if os.path.exists(path + '\\' + today_str + '出勤異常明細-非責任制.xlsx'):
        os.remove(path + '\\' + today_str + '出勤異常明細-非責任制.xlsx')
    wb1 = load_workbook(src_file)
    sheet1 = wb1.active
    wb2 = load_workbook(dl_file)
    sheet2 = wb2.active
    diff = [['日期', '班別', '工號', '姓名', '部門代碼', '部門', '出勤卡機狀態', 'PAD點名狀態']]
    list1 = []
    for i in range(1, sheet1.max_row):
        src = sheet1.cell(row=i + 1, column=1).value
        list1.append(src)
        src = sheet1.cell(row=i + 1, column=7).value
        list1.append(src)
        src = sheet1.cell(row=i + 1, column=4).value.capitalize()  # (capitalize工号首字母大写）
        list1.append(src)
        src = sheet1.cell(row=i + 1, column=5).value
        list1.append(src)
        src = sheet1.cell(row=i + 1, column=2).value
        list1.append(src)
        src = sheet1.cell(row=i + 1, column=3).value
        list1.append(src)
        src = sheet1.cell(row=i + 1, column=9).value
        list1.append(src)
    print('srcsk_list' + str(list1))
    list2 = []
    for i in range(1, sheet2.max_row):
        chk = sheet2.cell(row=i + 1, column=2).value.strftime('%Y/%m/%d')
        list2.append(chk)
        chk = sheet2.cell(row=i + 1, column=7).value
        list2.append(chk)
        chk = sheet2.cell(row=i + 1, column=8).value.capitalize()  # (capitalize工号首字母大写）
        list2.append(chk)
        chk = sheet2.cell(row=i + 1, column=9).value
        list2.append(chk)
        chk = sheet2.cell(row=i + 1, column=12).value
        list2.append(chk)
        chk = sheet2.cell(row=i + 1, column=13).value
        list2.append(chk)
        chk = sheet2.cell(row=i + 1, column=11).value
        list2.append(chk)
    print('dlddm_list' + str(list2))
    # 統計全部刷卡點名異常
    for i in range(2, len(list1), 7):
        if list1[i] in list2:  # 有刷卡有點名
            dmindex = list2.index(list1[i])
            if list2[dmindex + 4] in ['請假', '調休', '出差', '曠工', '24_無班']:
                # if (list2[dmindex + 3] == '請假' or list2[dmindex + 3] == '調休' or list2[dmindex+3] == '出差' or list2[dmindex+3] == '休假' or list2[dmindex+3] == '曠工'):
                diff_s = []
                diff_s.append(list1[i - 2])
                diff_s.append(list1[i - 1])
                diff_s.append(list1[i])
                diff_s.append(list1[i + 1])
                diff_s.append(list1[i + 2])
                diff_s.append(list1[i + 3])
                diff_s.append(list1[i + 4])
                diff_s.append(list2[dmindex + 4])
                diff.append(diff_s)
        elif list1[i] not in list2:  # 有刷卡無點名:
            diff_s = []
            diff_s.append(list1[i - 2])
            diff_s.append(list1[i - 1])
            diff_s.append(list1[i])
            diff_s.append(list1[i + 1])
            diff_s.append(list1[i + 2])
            diff_s.append(list1[i + 3])
            diff_s.append(list1[i + 4])
            diff_s.append('未點名')
            diff.append(diff_s)
    for j in range(2, len(list2), 7):
        if list2[j] not in list1:  # 有點名無刷卡
            if list2[j + 4] in ['有到', '銷假', '遲到']:
                # if (list2[j + 3] == '有到' or list2[j + 3] == '銷假' or list2[j + 3] == '遲到'):
                diff_s = []
                diff_s.append(list2[j - 2])
                diff_s.append(list2[j - 1])
                diff_s.append(list2[j])
                diff_s.append(list2[j + 1])
                diff_s.append(list2[j + 2])
                diff_s.append(list2[j + 3])
                diff_s.append('未刷卡')
                diff_s.append(list2[j + 4])
                diff.append(diff_s)
    print('fzrz_Diff:')
    print(diff)
    #   将数据写入excel
    wb3 = Workbook()
    sheet = wb3.active
    sheet.title = '刷卡點名出勤異常明細-All'
    for i in diff:
        sheet.append(i)
    wb3.save(path + '\\' + today_str + '刷卡點名異常明細-All.xlsx')
    # # 剔除責任制
    # dfskdmyc = DataFrame(read_excel(path + '\\' + today_str + '刷卡點名異常明細-All.xlsx'))
    # dfzrz = DataFrame(read_excel(path + '\\' + '責任制名單.xlsx'))
    # dfskdmycgh = dfskdmyc['工號'].tolist()
    # dfzrzgh = dfzrz['工號'].tolist()
    # print(dfzrzgh)
    # list3 = []
    # for i in dfskdmycgh:
    #     if i not in dfzrzgh:
    #         print(i)
    #         list3.append(i)
    # print(list3)
    # print(len(list3))
    # dfskdmyc = dfskdmyc.loc[dfskdmyc['工號'].isin(list3), ['部門代碼', '部門', '工號', '姓名', '班別', '出勤卡機狀態', 'PAD點名狀態']]
    # print(dfskdmyc)
    # dfskdmyc.to_excel(path + '\\' + today_str + '出勤異常明細-非責任制.xlsx', index=False)  # 写入excel文件，不包含索引数据

    # 從全部人力剔除全部責任制(非責任制人力)
    dfrenli = DataFrame(read_excel(path + '\\' + 'merged_all_renli.xlsx'))
    dfzrzrl = DataFrame(read_excel(path + '\\' + '全部責任制名單.xlsx'))
    dfrenligh = dfrenli['工號'].tolist()
    dfzrzrlgh = dfzrzrl['員工工號'].tolist()
    list4 = []
    for i in dfrenligh:
        if i not in dfzrzrlgh:
            list4.append(i)
    dffeizrzrl = dfrenli.loc[dfrenli['工號'].isin(list4), ['部門編號', '部門名稱', '工號', '姓名', '班別']]
    print('dffeizrzrl:')
    print(dffeizrzrl)
    dffeizrzrl.to_excel(path + '\\' + today_str + '人力-非責任制.xlsx', index=False)  # 写入excel文件，不包含索引数据

    # 剔除非三厰人員-（非責任制人員出勤）
    dfskdmyc = DataFrame(read_excel(path + '\\' + today_str + '刷卡點名異常明細-All.xlsx'))
    dffzrzrl = DataFrame(read_excel(path + '\\' + today_str + '人力-非責任制.xlsx'))
    dfskdmycgh = dfskdmyc['工號'].tolist()
    dffzrzrlgh = dffzrzrl['工號'].tolist()
    list3 = []
    for i in dfskdmycgh:
        if i in dffzrzrlgh:
            list3.append(i)
    dfskdmyc = dfskdmyc.loc[dfskdmyc['工號'].isin(list3), ['日期', '班別', '工號', '姓名', '部門代碼', '部門', '出勤卡機狀態', 'PAD點名狀態']]

    # 删除重复人员
    dfskdmyc.duplicated('工號')
    dupli_data = dfskdmyc.duplicated('工號').sum()
    print(f'有{dupli_data}个重复行')
    del_dupli = dfskdmyc[dfskdmyc.duplicated('工號', keep='last', )]  # 查看删除的重复的行
    print(f'删除的重复行为\n{del_dupli}')
    dfskdmyc.drop_duplicates('工號', keep='last', inplace=True)  # 删除最后一个重复行前面的重复行
    print('dfskdmyc:\n')
    print(dfskdmyc)
    dfskdmyc.to_excel(path + '\\' + today_str + '出勤異常明細-非責任制.xlsx', index=False)  # 写入excel文件，不包含索引数据


def excel_to_html(excel_f, html_f):
    if os.path.exists(html_f):
        os.remove(html_f)
    xd = ExcelFile(excel_f)
    df = xd.parse()
    with open(html_f, 'w', 'utf-8') as html_file:
        html_file.write(df.to_html(header=True, index=False))


def send_mail(content_fzrz, content_zrz, attach1, attach2):
    today = date.today()
    today_str = today.strftime('%Y%m%d')
    pythoncom.CoInitialize()
    ol_format_html = 2
    # OLFormatPlain = 1
    # olFormatRichText = 3
    # olFormatUnspecified = 0
    ol_mail_item = 0x0
    obj = Dispatch("Outlook.Application")
    send_account = None
    # 选择要使用的邮箱账户
    set_send_account = 'Apple_Zhang@intra.pegatroncorp.com'
    print('-------------------------------------')
    print('设定的邮件发件人与outlook读出来的必须一致！')
    print('添加outlook账号时，请将账号首字母大写，如 Carolyn_Yu@pegatroncorp.com！')
    print('-------------------------------------')
    print('程式设定的邮件发件人是: %s' % set_send_account)
    print('Outlook读出来的账号如下:')
    for account in obj.Session.Accounts:
        print(account)
        if account.DisplayName == set_send_account:  # 需与读出来的一致，否则报错
            send_account = account
            break
    new_mail = obj.CreateItem(ol_mail_item)

    new_mail._oleobj_.Invoke(*(64209, 0, 8, 0), send_account)  # 指定发件人

    new_mail.Subject = '自動郵件： ' + today_str + '出勤异常预警!'
    new_mail.BodyFormat = ol_format_html
    # new_mail.HTMLBody = "<h1>I am a title</h1><p>I am a paragraph</p>"
    # new_mail.HTMLBody = """
    #              <p>Hello All!&nbsp;</p>
    #                   未点名人员如下，请确认：
    #              """ + open(file_folder + '\\' + today_str + '未點名人員名單.html', "r", encoding='utf-8').read()
    new_mail.HTMLBody = """
                 <p>Hello All!&nbsp;</p>                      
                 """ + content_fzrz + open(file_folder + '\\' + today_str + '出勤異常明細-非責任制.html', "r",
                                           encoding='utf-8').read() + """
                 <p>_&nbsp;</p>                      
                 """ + content_zrz + open(file_folder + '\\' + today_str + '未點名人員名單-責任制.html', "r",
                                          encoding='utf-8').read()
    new_mail.To = EX_MAILTO

    # carbon copies and attachments (optional)
    new_mail.CC = EX_MAILCC
    # new_mail.BCC = "Hong-ze_Wang@pegatroncorp.com"
    new_mail.Attachments.Add(attach1)
    new_mail.Attachments.Add(attach2)

    # open up in a new window and allow review before send
    # new_mail.display()

    # or just use this instead of .display() if you want to send immediately
    new_mail.Send()
    pythoncom.CoUninitialize()  # 多线程需要pythoncom初始及关闭，否则报错
    print('Mail sent successfully!')
    #  如何禁用系统弹窗: OutLook选项——信任中心——信任中心设置——编程访问——从不向我发出可疑活动警告


def read_mail_addr(mail_adr, mail_cc):
    while os.path.exists(MAIL_PATH):
        break
    else:
        print("No Mail address file, please insert 'MailAddressListTemplet.xls' to " + MAIL_PATH)
        sleep(1)
        read_mail_addr(mail_adr, mail_cc)
    wb = load_workbook(MAIL_PATH)
    # ws = wb['Sheet1']
    ws = wb.active
    if ws.max_row == 0 or ws.max_row == 1:
        print('There have not mail address data, please insert!')
        sleep(1)
        read_mail_addr(mail_adr, mail_cc)
    else:
        regex = compile(r'([A-Za-z0-9]+[-._])*[A-Za-z0-9]+@[A-Za-z0-9-]+(\.[A-Z|a-z]{2,})+')
        for i in range(2, ws.max_row + 1):  # 打印一列表格的内容
            if ws.max_column == 1:
                if ws.cell(i, 1).value is not None:
                    print(type(ws.cell(i, 1).value))
                    if fullmatch(regex, ws.cell(i, 1).value):
                        mail_adr = mail_adr + ws.cell(i, 1).value + ';'
                        mail_cc = ''
                    else:
                        print('无效的email地址' + ws.cell(i, 1).value)
                        sleep(1)
                        read_mail_addr(mail_adr, mail_cc)
            else:
                if ws.cell(i, 1).value is not None:
                    if fullmatch(regex, ws.cell(i, 1).value):
                        mail_adr = mail_adr + ws.cell(i, 1).value + ';'
                    else:
                        print('无效的email地址' + ws.cell(i, 1).value)
                        sleep(1)
                        read_mail_addr(mail_adr, mail_cc)
                if ws.cell(i, 2).value is not None:
                    if fullmatch(regex, ws.cell(i, 2).value):
                        mail_cc = mail_cc + ws.cell(i, 2).value + ';'
                    else:
                        print('无效的email地址' + str(ws.cell(i, 2).value))
                        sleep(1)
                        read_mail_addr(mail_adr, mail_cc)
    return mail_adr, mail_cc


def check_webdriver(browser=browser_driver):
    while os.path.exists(browser):
        print('瀏覽器driver文件路徑：' + browser + ', browser driver file is OK!')
        break
    else:
        print("No browser driver file, please insert browser driver file to " + file_folder)
        sleep(1)
        check_webdriver()


def job():
    try:
        del_serial_file(file_folder, '刷卡點名異常明細-All', 'xlsx')
        del_serial_file(file_folder, '人力-非責任制', 'xlsx')
        del_serial_file(file_folder, '出勤異常明細-非責任制', 'html')
        del_serial_file(file_folder, '出勤異常明細-非責任制', 'xlsx')
        del_serial_file(file_folder, '未點名人員名單-責任制', 'html')
        del_serial_file(file_folder, '未點名人員名單-責任制', 'xlsx')
        # msg_alert_psgui()  # 线程里调用pysimplegui报错 main thread not in main loop
        print("I'm running on schedule!")
        print("Python自动流程即将开始，请勿动鼠标与键盘！~")
        today = date.today()
        today_str = today.strftime('%Y%m%d')
        print(today_str)
        if_exist_foler(file_folder)
        dl_data()  # download 人力及点名
        dl_sk_data()  # download 刷卡信息
        # copy_serial_file(dl_di, 'Data', 'xlsx')
        # copy_serial_file(dl_di, 'Attendance', 'xlsx')
        merge_excel(dl_di, file_folder, 'Data', 'xlsx', 'merged_all_renli.xlsx')  # 汇整人力数据
        dianm_shaix(dl_di, file_folder, 'merged_dianm.xlsx')  # 汇整点名数据
        shuaka_shaix(dl_di, file_folder, 'merged_shuaka.xlsx')  # 汇整刷卡数据
        # 筛选责任制未點名异常数据
        compare_zrzrl_dm(file_folder + '\\' + '需統計出勤責任制名單.xlsx', file_folder + '\\' + 'merged_dianm.xlsx', file_folder)
        excel_to_html(file_folder + '\\' + today_str + '未點名人員名單-責任制.xlsx',
                      file_folder + '\\' + today_str + '未點名人員名單-責任制.html')
        # 筛选非责任制刷卡点名异常数据
        comp_sk_dm(file_folder + '\\' + 'merged_shuaka.xlsx', file_folder + '\\' + 'merged_dianm.xlsx',
                   file_folder)
        excel_to_html(file_folder + '\\' + today_str + '出勤異常明細-非責任制.xlsx',
                      file_folder + '\\' + today_str + '出勤異常明細-非責任制.html')
        # fzrz判斷是否有異常
        wb1 = load_workbook(file_folder + '\\' + today_str + '出勤異常明細-非責任制.xlsx')
        sheet1 = wb1.active
        if sheet1['a2'].value is not None:
            content_fzrz = '截止目前（非責任制）出勤異常明細如下，請確認!'
        else:
            content_fzrz = '截止目前非責任制無出勤異常!'

        # zrz判斷是否有異常
        wb1 = load_workbook(file_folder + '\\' + today_str + '未點名人員名單-責任制.xlsx')
        sheet1 = wb1.active
        if sheet1['a2'].value is not None:
            content_zrz = '截止目前（責任制）未點名人員明細如下，請確認!'
        else:
            content_zrz = '截止目前責任制人員都已經點名啦!'
        send_mail(content_fzrz, content_zrz, file_folder + '\\' + today_str + '出勤異常明細-非責任制.xlsx',
                  file_folder + '\\' + today_str + '未點名人員名單-責任制.xlsx')
        # complete_alert_psgui()  # 线程里调用pysimplegui报错main thread not in main loop
        print('出勤統計完成.')
    except Exception as job_er:
        print('job_erorr: %s' % job_er)
        # print('请将{}窗口激活并最大化！'.format(win_name))
        job()


def run_threaded():
    job_thread = Thread(target=job)
    job_thread.setDaemon(True)  # 创建多线程 设置以保护模式启动，即主线程运行结束，子线程也停止运行
    job_thread.start()


def unlock_scr():
    while True:
        pyautogui.FAILSAFE = False  # 当为True或不加时鼠标放到角落会报错退出
        print('unlock scr!!!')
        pyautogui.press('volumedown')
        sleep(1)
        pyautogui.press('volumeup')
        sleep(90)


def sch_input():
    try:
        print('请输入流程每天执行时间点(至少比當前時間晚1分鐘)，时间输入格式如08：06')
        tm = input()
        # every().day.at(tm).do(run_threaded)
        every().monday.at(tm).do(run_threaded)
        every().tuesday.at(tm).do(run_threaded)
        every().wednesday.at(tm).do(run_threaded)
        every().thursday.at(tm).do(run_threaded)
        every().friday.at(tm).do(run_threaded)
        # every().saturday.at(tm).do(run_threaded)
        # every().sunday.at(tm).do(run_threaded)
    except Exception as input_er:
        print('sch_input error: %s' % input_er)
        print('输入格式有误，请切换到英文状态重新输入!')
        sch_input()


# 提示信息
print('*' * 100 + '\n' + '*' * 6 + ' 注意事项：')
print('*' * 6 + " 请打开Outlook! 请不要锁屏（可以关闭屏幕）!")
print('*' * 6 + " 請將'chromedriver.exe'文件放到%s文件夹下!" % file_folder)
print('*' * 6 + " 請將郵件地址文件'MailAddressListTemplet.xlsx'放到%s文件夹下!" % file_folder)
print('*' * 6 + " 請將責任制人力文件'全部責任制名單.xlsx'和'需統計出勤責任制名單'放到%s文件夹下!" % file_folder)
print('*' * 6 + " 請確保設定的自動下載路徑為：%s" % dl_di)
print('*' * 6 + ' 出勤及點名源數據下載路徑：' + dl_di)
print('*' * 6 + ' 責任制異常數據文件：' + file_folder + '\\' + '未點名人員名單-責任制.xlsx')
print('*' * 6 + ' 非責任制異常數據文件：' + file_folder + '\\' + '出勤異常明細-非責任制.xlsx' + '\n' + '*' * 100)
check_webdriver()  # 瀏覽器driver文件是否存在
print('郵件地址文件路徑：' + MAIL_PATH + ', Mail address file is OK!')
mail_list = read_mail_addr(EX_MAILTO, EX_MAILCC)
EX_MAILTO = mail_list[0]  # 读取收件者邮件地址
EX_MAILCC = mail_list[1]  # 读取副本邮件地址
print('收件者: ' + EX_MAILTO)  # 收件者
print('副 本: ' + EX_MAILCC)  # 副本
sleep(5)

if __name__ == '__main__':
    dept_dianm = dept_input(dept_dianm, '點名結果', '(名碩)第五事業處-製造二處-支援部-製三課、(名碩)第五事業處-製造三處-三廠')
    dept_renli = dept_input(dept_renli, '人力清冊', '(名碩)第五事業處-製造二處-支援部-製三課、(名碩)第五事業處-製造三處-三廠')
    dept_shuaka = dept_input(dept_shuaka, '非責任制出勤記錄查詢', 'ALL')
    sch_input()
    print('Waiting for time to run schedule!')
    # every().day.at('15:39').do(run_threaded)
    # every().day.at('10:00').do(run_threaded)
    # # main thread is not in main loop报错，daemon为True，就是我们平常理解的后台线程，用Ctrl-C关闭程序，所有后台线程都会被自动关闭。
    # 如果daemon属性是False，线程不会随主线程的结束而结束，这时如果线程访问主线程的资源，就会出错。
    mail_cfm_win = Thread(target=mail_cfm_win, args=('Microsoft Outlook',))
    mail_cfm_win.setDaemon(True)  # 创建多线程 设置以保护模式启动，即主线程运行结束，子线程也停止运行
    mail_cfm_win.start()
    t_unlock_scr = Thread(target=unlock_scr)
    t_unlock_scr.setDaemon(True)  # 创建多线程 设置以保护模式启动，即主线程运行结束，子线程也停止运行
    t_unlock_scr.start()
    while True:
        try:
            run_pending()  # 检查上面的任务部署情况，如果任务已经准备就绪，就去启动执行。
            # run_all()  # 立即执行
            sleep(10)  # 让程序按10秒来检查
        except Exception as e:
            print('main error: %s' % e)
