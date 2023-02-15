import openpyxl
import pprint


def read_dian_ming_biao(dmbiao):
    wb = openpyxl.load_workbook(dmbiao)
    sheet = wb['Sheet1']
    sheet_copy = wb.copy_worksheet(sheet)
    rowmax = str(sheet_copy.max_row)
    date = str(sheet_copy['B' + rowmax].value)
    print(date)
    # rename 班别为白班/夜班，区分为DL/IDL
    for i in range(2, sheet_copy.max_row + 1):
        if sheet_copy['G' + str(i)].value[0:2] == "白班":
            sheet_copy['G' + str(i)].value = "白班"
        else:
            sheet_copy['G' + str(i)].value = "夜班"
        if sheet_copy['J' + str(i)].value == 1:
            sheet_copy['J' + str(i)] = 'DL'
        elif sheet_copy['J' + str(i)].value == 2:
            sheet_copy['J' + str(i)] = 'DL'
        else:
            sheet_copy['J' + str(i)] = 'IDL'
    wb.save(dmbiao)

    # data = {}
    data = {'B5E3400M0E': {'DL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}},
                           'IDL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}}
                           },
            'B5E3410M0E': {'DL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}},
                           'IDL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}}
                           },
            'B5E3420M0E': {'DL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}},
                           'IDL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}}
                           },
            'B5E3430M0E': {'DL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}},
                           'IDL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}}
                           },
            'B5B0B3CM0E': {'DL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}},
                           'IDL': {'白班': {'有到': 0, '請假': 0}, '夜班': {'有到': 0, '請假': 0}}
                           },
            }
    # 统计人力
    for i in range(2, sheet_copy.max_row + 1):
        bumen = sheet_copy['L' + str(i)].value
        zhideng = sheet_copy['J' + str(i)].value
        shift = sheet_copy['G' + str(i)].value
        chuqing = sheet_copy['K' + str(i)].value
        # name = sheet_copy['I' + str(i)].value
        # data.setdefault(bumen, {})
        # data[bumen].setdefault(zhideng, {})
        # data[bumen][zhideng].setdefault(shift, {})
        data[bumen][zhideng][shift].setdefault(chuqing, 0)
        data[bumen][zhideng][shift][chuqing] += 1
        # data[bumen][zhideng][shift].setdefault(chuqing,{'mingzi':0})
        # data[bumen][zhideng][shift][chuqing]['mingzi'] +=1
        # print(data[bumen][zhideng][shift][chuqing]['mingzi'])
    pprint.pprint(data)
    # print(pprint.pformat(data))
    '''
    # Open a new text file and write the content of data to it.
    print('Writing results...')
    resultFile = open('renli2.py', 'w')
    resultFile.write('allData= ' + pprint.pformat(data))
    resultFile.close()
    print('Done')
    '''

    # print(data['B5E3410M0E']['DL']['白班']['有到'])

    wb2 = openpyxl.load_workbook("人員統計 3400.xlsx")
    sheet21 = wb2.active
    allsheet = wb2.sheetnames
    for sheetName in allsheet:
        if sheetName == date[0:10]:
            wb2.remove(wb2[sheetName])
    sheet22 = wb2.copy_worksheet(sheet21)
    sheet22.title = date[0:10]

    # clear old data
    for i in sheet22.iter_rows(min_row=19, max_row=23, min_col=4, max_col=19):
        for j in i:
            j.value = ''


    # write data to excel
    def write_to_tongji(bum, zhid, banb, youd, qingj, yingd, shid, weid, yuany):
        # B5E3400M0E DL 白班
        sheet22[shid].value = data[bum][zhid][banb][youd]
        sheet22[weid].value = data[bum][zhid][banb][qingj]
        if sheet22[weid].value == 0:
            sheet22[yingd].value = sheet22[shid].value
        else:
            sheet22[yingd].value = sheet22[shid].value + sheet22[weid].value
            sheet22[yuany].value = '请假:家中有事'

    # B5E3400M0E DL 白班
    write_to_tongji('B5E3400M0E', 'DL', '白班', '有到', '請假', 'D19', 'E19', 'F19', 'G19')
    write_to_tongji('B5E3400M0E', 'DL', '夜班', '有到', '請假', 'H19', 'I19', 'J19', 'K19')
    write_to_tongji('B5E3400M0E', 'IDL', '白班', '有到', '請假', 'L19', 'M19', 'N19', 'O19')
    write_to_tongji('B5E3400M0E', 'IDL', '夜班', '有到', '請假', 'P19', 'Q19', 'R19', 'S19')
    write_to_tongji('B5E3410M0E', 'DL', '白班', '有到', '請假', 'D20', 'E20', 'F20', 'G20')
    write_to_tongji('B5E3410M0E', 'DL', '夜班', '有到', '請假', 'H20', 'I20', 'J20', 'K20')
    write_to_tongji('B5E3410M0E', 'IDL', '白班', '有到', '請假', 'L20', 'M20', 'N20', 'O20')
    write_to_tongji('B5E3410M0E', 'IDL', '夜班', '有到', '請假', 'P20', 'Q20', 'R20', 'S20')
    write_to_tongji('B5E3420M0E', 'DL', '白班', '有到', '請假', 'D21', 'E21', 'F21', 'G21')
    write_to_tongji('B5E3420M0E', 'DL', '夜班', '有到', '請假', 'H21', 'I21', 'J21', 'K21')
    write_to_tongji('B5E3420M0E', 'IDL', '白班', '有到', '請假', 'L21', 'M21', 'N21', 'O21')
    write_to_tongji('B5E3420M0E', 'IDL', '夜班', '有到', '請假', 'P21', 'Q21', 'R21', 'S21')
    write_to_tongji('B5E3430M0E', 'DL', '白班', '有到', '請假', 'D22', 'E22', 'F22', 'G22')
    write_to_tongji('B5E3430M0E', 'DL', '夜班', '有到', '請假', 'H22', 'I22', 'J22', 'K22')
    write_to_tongji('B5E3430M0E', 'IDL', '白班', '有到', '請假', 'L22', 'M22', 'N22', 'O22')
    write_to_tongji('B5E3430M0E', 'IDL', '夜班', '有到', '請假', 'P22', 'Q22', 'R22', 'S22')
    write_to_tongji('B5B0B3CM0E', 'DL', '白班', '有到', '請假', 'D23', 'E23', 'F23', 'G23')
    write_to_tongji('B5B0B3CM0E', 'DL', '夜班', '有到', '請假', 'H23', 'I23', 'J23', 'K23')
    write_to_tongji('B5B0B3CM0E', 'IDL', '白班', '有到', '請假', 'L23', 'M23', 'N23', 'O23')
    write_to_tongji('B5B0B3CM0E', 'IDL', '夜班', '有到', '請假', 'P23', 'Q23', 'R23', 'S23')

    # replace 0 to ''
    for i in sheet22.iter_rows(min_row=19, max_row=23, min_col=4, max_col=19):
        for j in i:
            if j.value == 0:
                j.value = ''

    wb2.save("人員統計 3400.xlsx")


read_dian_ming_biao('點名歷史記錄_20220120.xlsx')
