# _*_ coding: utf-8 _*_
# @Time   : 2022/3/10 下午 01:08
# @Author : 未来战士yuyu!!
# @FileName : 人力统计.py
# @Software : PyCharm
# @Blog     : http://blog.csdn.net/u010105243/article/
import openpyxl
import numpy as np
import pandas as pd


def read_dian_ming_biao(dmbiao):
    df = pd.DataFrame(pd.read_excel(dmbiao, sheet_name='Sheet1'))
    # df_inner.loc[(df_inner['city'] == 'Beijing') & (df_inner['price'] >= 1000),'sign'] = 1
    # # df_inner['group'] = np.where(df_inner['price'] > 3000, 'high', 'low')
    # 新建列：班别为白班/夜班，区分为DL/IDL,異常原因为有到/請假
    df['區分S'] = np.where(df['區分'].isin(['3', '4', '5', '6', '7', '8']), 'IDL', 'DL')
    df['班別S'] = np.where(df['班別'].isin(['夜班四', '夜班六', '夜班八']), '夜班', '白班')
    df['異常原因S'] = np.where(df['異常原因'] == '請假', '請假', '有到')
    # print(df[['區分', '區分S', '班別', '班別S','異常原因','異常原因S']])
    '''
    # 方法一、按部门、职等、班别，出勤统计人数
    total_list = []
    for bm in df['支援部門代碼'].unique():
        for zhid in df['區分S'].unique():
            for banb in df['班別S'].unique():
                for chuq in df['異常原因S'].unique():
                    num = df.loc[(df['支援部門代碼'] == bm) & (df['區分S'] == zhid) & (df['班別S'] == banb) & (df['異常原因S'] == chuq),
                                 ['支援部門代碼', '區分', '區分S', '班別', '班別S', '異常原因','異常原因S', '姓名']].姓名.count()
                    total_list.append([bm, zhid, banb, chuq, num])
                    # print(bm, zhid, banb, chuq, df.loc[(df['支援部門代碼'] == bm) & (df['區分S'] == zhid) & (df['班別S'] == banb) & (df['異常原因S'] == chuq),
                    #                                    ['支援部門代碼', '區分', '區分S', '班別', '班別S', '異常原因','異常原因S', '姓名']].姓名.count())
    print(total_list)
    df3 = pd.DataFrame(total_list)
    print(df3)
    with pd.ExcelWriter(r'G:\PycharmProjects\pythonProject\manpower\result.xlsx') as writer:
        df3.to_excel(writer, sheet_name='匹配结果')
    '''
    # 方法二、按部门、职等、班别，出勤统计人数
    bm_list = ['B5E3400M0E', 'B5E3410M0E', 'B5E3420M0E', 'B5E3430M0E', 'B5B0B3CM0E']
    zhid_list = ['DL', 'IDL']
    banb_list = ['白班', '夜班']
    chuq_list = ['有到', '請假']
    # total_list = []
    sum_list = []
    for bm in bm_list:
        sub_list = []
        # sub_list.append(bm)
        for zhid in zhid_list:
            for banb in banb_list:
                for chuq in chuq_list:
                    num = df.loc[
                        (df['支援部門代碼'] == bm) & (df['區分S'] == zhid) & (df['班別S'] == banb) & (df['異常原因S'] == chuq),
                        ['支援部門代碼', '區分', '區分S', '班別', '班別S', '異常原因', '異常原因S', '姓名']].姓名.count()
                    sub_list.append(num)
                    print(bm, zhid, banb, chuq, num)
        # total_list.append(sub_list)
        list4 = [sub_list[0] + sub_list[1], sub_list[0], sub_list[1], '', sub_list[2] + sub_list[3], sub_list[2],
                 sub_list[3], '', sub_list[4] + sub_list[5], sub_list[4], sub_list[5], '', sub_list[6] + sub_list[7],
                 sub_list[6], sub_list[7], '']
        sum_list.append(list4)

    print(sum_list)

    wb = openpyxl.load_workbook(dmbiao)
    sheet = wb['Sheet1']
    sheet_copy = wb.copy_worksheet(sheet)
    rowmax = str(sheet_copy.max_row)
    date = str(sheet_copy['B' + rowmax].value)
    print(date)

    wb2 = openpyxl.load_workbook("人員統計 3400.xlsx")
    sheet21 = wb2.active
    allsheet = wb2.sheetnames
    for sheetName in allsheet:
        if sheetName == date[0:10]:
            wb2.remove(wb2[sheetName])
    sheet22 = wb2.copy_worksheet(sheet21)
    sheet22.title = date[0:10]

    # replace 0 to ''
    for i in sheet22.iter_rows(min_row=19, max_row=23, min_col=4, max_col=19):
        for j in i:
            if j.value == 0:
                j.value = ''

    # write data to excel
    print(len(sum_list[0]))
    # for i in sheet22.iter_rows(min_row=19, max_row=23, min_col=4, max_col=19):
    #     for j in i:
    #         for x in range(0,len(sum_list)):
    #             for y in range(0,len(sum_list[x])):
    #                 print(j)
    #                 print(sum_list[x][y])
    #                 j.value = sum_list[x][y]
    #                 print(j.value)

    for i in range(len(sum_list)):
        for j in range(0, len(sum_list[i])):
            sheet22.cell(i + 19, j + 4, sum_list[i][j])

    # 填写未到原因
    for i in range(0, 5):
        for j in range(0, 1):
            if sheet22.cell(i + 19, j + 6).value != 0:
                sheet22.cell(i + 19, j + 7).value = '请假'
            if sheet22.cell(i + 19, j + 10).value != 0:
                sheet22.cell(i + 19, j + 11).value = '请假'
            if sheet22.cell(i + 19, j + 14).value != 0:
                sheet22.cell(i + 19, j + 15).value = '请假'
            if sheet22.cell(i + 19, j + 18).value != 0:
                sheet22.cell(i + 19, j + 19).value = '请假'

    # replace 0 to ''
    for i in sheet22.iter_rows(min_row=19, max_row=23, min_col=4, max_col=19):
        for j in i:
            if j.value == 0:
                j.value = ''

    wb2.save("人員統計 3400.xlsx")


read_dian_ming_biao('20220301Attendance.xlsx點名.xlsx')
